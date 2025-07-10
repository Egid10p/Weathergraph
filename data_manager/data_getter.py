# Copyright 2025 Egidio Pulicanò
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

import os
from typing import Optional
import openpyxl
import openpyxl.utils
import openpyxl.utils.exceptions
import openpyxl.worksheet
import openpyxl.worksheet.worksheet
from utils.file_initializer import ensure_file_exists


class DataGetter:
    """
    Provides structured and flexible access to data within Excel (.xlsx) spreadsheets using openpyxl.

    This class enables reading tabular data from specific sheets, either partially (top or bottom N rows)
    or entirely, returning it in dictionary format with headers mapped to row values.
    """

    def __init__(self, path: str):
        """
        Initializes the DataGetter instance with the specified file path.

        Args:
            path (str): Absolute or relative path to the Excel file.
            
        Raises:
            ValueError: If the specified file does not exist or is not a valid Excel workbook.
        """
        
        if not os.path.isfile(path):
           raise ValueError(f"File does not exist: {path!r}")
        self.path = path
        self.workbook = None

    def open_file(self) -> None:
        """
        Loads the Excel workbook into memory if it has not already been loaded.

        Raises:
            FileNotFoundError: If the specified file does not exist.
            ValueError: If the file cannot be opened as a valid Excel workbook.
        """
        if self.workbook is None:
            try:
                self.workbook = openpyxl.load_workbook(self.path)
            except FileNotFoundError:
                raise FileNotFoundError(f"The file does not exist: {self.path}")
            except (openpyxl.utils.exceptions.InvalidFileException, OSError) as e:
                raise ValueError(f"Error opening the file: {e}")

    def close_file(self) -> None:
        """
        Closes the currently loaded workbook and clears the internal reference.
        Useful for explicitly releasing file handles.
        """
        if self.workbook:
            self.workbook.close()
            self.workbook = None

    def get_sheet(self, sheet_name: Optional[str] = None) -> openpyxl.worksheet.worksheet.Worksheet:
        """
        Retrieves a worksheet by name or returns the active sheet if no name is provided.

        Args:
            sheet_name (Optional[str]): Name of the worksheet to retrieve. If None, uses the active sheet.

        Returns:
            openpyxl.worksheet.worksheet.Worksheet: Worksheet object.

        Raises:
            RuntimeError: If no workbook has been loaded prior to access.
            ValueError: If the specified sheet name does not exist in the workbook.
        """
        if not self.workbook:
            raise RuntimeError("Workbook is not loaded.")
        sheet = self.workbook[sheet_name] if sheet_name else self.workbook.active
        if sheet is None:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook.")
        return sheet

    def _get_headers(self, sheet: Optional[str] = None) -> list[str]:
        """
        Extracts column headers from the first row of the specified sheet.

        Args:
            sheet (Optional[str]): Name of the sheet to target. Uses active sheet if None.

        Returns:
            list[str]: A list of header names.
            
        Raises:
            ValueError: If the sheet object is None or if the first row is empty.
        """
        sheet_obj = self.get_sheet(sheet)
        if sheet_obj is None:
            raise ValueError("Sheet object is None.")
        return [str(cell.value) if cell.value is not None else "" for cell in sheet_obj[1]]

    def _get_top_rows(self, n_rows: int, sheet: Optional[str] = None) -> list[tuple]:
        """
        Retrieves the first `n_rows` data rows (after the header) from the given sheet.

        Args:
            n_rows (int): Number of top rows to return (excluding header).
            sheet (Optional[str]): Sheet to extract rows from.

        Returns:
            list[tuple]: List of tuples representing row values.

        Raises:
            ValueError: If `n_rows` is less than 1.
        """
        if n_rows < 1:
            raise ValueError("n_rows must be a positive integer.")
        sheet_obj = self.get_sheet(sheet)
        return [
            row for row in sheet_obj.iter_rows(
                min_row=2,
                max_row=min(n_rows + 1, sheet_obj.max_row),
                values_only=True
            )
        ]

    def _get_bottom_rows(self, n_rows: int, sheet: Optional[str] = None) -> list[tuple]:
        """
        Retrieves the last `n_rows` data rows from the sheet.

        Args:
            n_rows (int): Number of bottom rows to return.
            sheet (Optional[str]): Sheet to extract rows from.

        Returns:
            list[tuple]: List of tuples representing the bottom rows.

        Raises:
            ValueError: If `n_rows` is less than 1.
        """
        if n_rows < 1:
            raise ValueError("n_rows must be a positive integer.")
        sheet_obj = self.get_sheet(sheet)
        min_row = max(2, sheet_obj.max_row - n_rows + 1)
        return [
            row for row in sheet_obj.iter_rows(
                min_row=min_row,
                max_row=sheet_obj.max_row,
                values_only=True
            )
        ]

    @staticmethod
    def _rows_to_dicts(rows: list[tuple], headers: list[str]) -> list[dict]:
        """
        Converts a list of row tuples into dictionaries using the provided headers.

        Args:
            rows (list[tuple]): List of data rows as tuples.
            headers (list[str]): Corresponding header names.

        Returns:
            list[dict]: List of dictionaries mapping headers to row values.

        Raises:
            ValueError: If any row length differs from the number of headers.
        """
        data = []
        for row in rows:
            if len(row) != len(headers):
                raise ValueError("Row length does not match header length.")
            data.append(dict(zip(headers, row)))
        return data
    
    def get_data(self, n_rows: int, from_top: bool = True, sheet: Optional[str] = None) -> list[dict]:
        """
        Retrieves a specific number of rows from the top or bottom of the sheet as dictionaries.

        Args:
            n_rows (int): Number of rows to retrieve.
            from_top (bool): Whether to fetch from the top (True) or bottom (False).
            sheet (Optional[str]): Sheet name to read from.

        Returns:
            list[dict]: List of dictionaries containing the row data.
        """
        headers = self._get_headers(sheet)
        if from_top:
            rows = self._get_top_rows(n_rows, sheet)
        else:
            rows = self._get_bottom_rows(n_rows, sheet)
        return self._rows_to_dicts(rows, headers)

    def get_all_data(self, sheet: Optional[str] = None) -> list[dict]:
        """
        Retrieves all available data rows (excluding header) from the sheet as dictionaries.

        Args:
            sheet (Optional[str]): Sheet to read from. Defaults to active sheet if not provided.

        Returns:
            list[dict]: Complete dataset mapped as dictionaries.
        """
        headers = self._get_headers(sheet)
        sheet_obj = self.get_sheet(sheet)
        rows = self._get_top_rows(sheet_obj.max_row - 1, sheet)
        return self._rows_to_dicts(rows, headers)
    
    def __enter__(self) -> "DataGetter":
        """
        Context manager entry point to open the workbook file.

        Returns:
            DataGetter: The current instance of DataGetter.
        """
        self.open_file()
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb) -> None:
        """
        Ensures the workbook is closed when the DataGetter instance is deleted.
        """
        self.close_file()