# Copyright 2025 Egidio Pulicanò
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

import os
from typing import Optional, Iterable, Sequence, Any
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
import openpyxl.utils.exceptions
from utils.file_initializer import ensure_file_exists


class DataSaver:
    """
    Append rows of structured data to an Excel (.xlsx) workbook using openpyxl.

    This utility ensures:
      - The target directory exists (created if necessary).
      - The file path ends with '.xlsx'.
      - The workbook is loaded from disk or created new.
      - The specified worksheet is selected or created.
      - Each row is validated and converted to list[Any] before insertion.
      - I/O errors during save operations raise clear exceptions.

    Attributes:
        file_path (str): Absolute path to the .xlsx file.
        sheet_name (Optional[str]): Name of the target worksheet.
        wb (Workbook): The loaded or created workbook instance.
        ws (Worksheet): The active or newly created worksheet.
    """

    def __init__(self, file_path: str, sheet_name: Optional[str] = None) -> None:
        """
        Initializes a DataSaver instance by ensuring the target Excel file exists,
        loading the workbook, and selecting or creating the desired worksheet.

        Args:
            file_path (str): Path to the Excel (.xlsx) file.
            sheet_name (Optional[str]): Name of the worksheet to use or create.
                                        Defaults to the active sheet if omitted.

        Raises:
            ValueError: If the workbook cannot be opened or if worksheet selection fails.
        """

        ensure_file_exists(file_path)
        
        self.file_path = file_path
        self.sheet_name = sheet_name

        # Load existing workbook or create a new one
        try:
            self.wb: Workbook = openpyxl.load_workbook(file_path)
        except Exception as e:
            raise ValueError(f"Cannot open workbook {file_path!r}: {e}")

        if sheet_name:
            if sheet_name in self.wb.sheetnames:
                ws = self.wb[sheet_name]
            else:
                ws = self.wb.create_sheet(sheet_name)
        else:
            ws = self.wb.active

        if not isinstance(ws, Worksheet):
            raise ValueError(f"Worksheet selection failed: {sheet_name!r}")
        self.ws: Worksheet = ws

    def save_rows(self, rows: Iterable[Sequence[Any]]) -> None:
        """
        Validate and append multiple rows, then persist the workbook.

        Args:
            rows (Iterable[Sequence[Any]]): An iterable of row sequences,
                each sequence representing a row of cell values.

        Raises:
            ValueError: If any row is not a non-empty list or tuple.
            OSError:    If saving the workbook to disk fails.
        """
        for row in rows:
            self._validate_row(row)
            self.ws.append(list(row))
        self._save_workbook()

    def save_data(self, row: Sequence[Any]) -> None:
        """
        Validate and append a single row, then persist the workbook.

        Args:
            row (Sequence[Any]): A sequence of values for one row.

        Raises:
            ValueError: If row is not a non-empty list or tuple.
            OSError:    If saving the workbook to disk fails.
        """
        self._validate_row(row)
        self.ws.append(list(row))
        self._save_workbook()

    def _validate_row(self, row: Sequence[Any]) -> None:
        """
        Ensure that a row is a non-empty list or tuple.

        Args:
            row (Sequence[Any]): The row to validate.

        Raises:
            ValueError: If row is not a list/tuple or is empty.
        """
        if not isinstance(row, (list, tuple)) or len(row) == 0:
            raise ValueError(
                f"Row must be a non-empty list or tuple, got {type(row).__name__}"
            )

    def _save_workbook(self) -> None:
        """
        Save the workbook to disk at the configured file path.

        Raises:
            OSError: If writing to disk fails.
        """
        try:
            self.wb.save(self.file_path)
        except OSError as e:
            raise OSError(f"Failed to save workbook {self.file_path!r}: {e}")

    def close(self) -> None:
        """
        Close the workbook and release file handles.

        The DataSaver instance should be discarded or reinitialized
        after calling this method.
        """
        try:
            self.wb.close()
        except Exception:
            pass

    def __enter__(self) -> "DataSaver":
        """
        Enter the runtime context related to this object.

        Returns:
            DataSaver: The DataSaver instance itself.
        """
        return self

    def __exit__(self, exc_type, exc_val, exc_tb) -> None:
        """
        Exit the runtime context and close the workbook.

        Args:
            exc_type: Exception type, if raised.
            exc_val: Exception value, if raised.
            exc_tb: Exception traceback, if raised.
        """
        self.close()
